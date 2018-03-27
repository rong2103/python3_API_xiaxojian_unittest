# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import os

class MyExcel:
    def __init__(self,filename,logger):
        self.wb = load_workbook(filename)
        self.logger = logger

    #读取某个cell值
    def get_init_variables(self,sheetname):
        sh = self.wb.get_sheet_by_name(sheetname)
        init_datas = {}
        for index in  range(2,sh.max_row+1):
            key = sh.cell(index,1).value
            init_datas[key] = str(sh.cell(index,2).value)
        init_datas["${phone}"] = str(int(init_datas["${init_phone}"]) + 1)
        init_datas["${noReg_phone}"] = str(int(init_datas["${init_phone}"]) + 2)
        self.logger.info("get_init_variables函数获取到的初始化变量为:{0}".format(init_datas))
        return init_datas

    # 读取excel所有行数据
    def getallrow(self,sheetname):
        sh = self.wb.get_sheet_by_name(sheetname)
        one_test_data = []
        dict1 = {}
        cols = sh.max_column
        rows = sh.max_row
        init_datas = self.get_init_variables("variables")
        for i in range(2,rows+1):
            for j in range(1, cols + 1):
                k = sh.cell(1, j).value
                v = str(sh.cell(i, j).value)
                dict1[k] = v
                for key,value in init_datas.items():
                    # 如果请求数据为空时需要做处理
                    if dict1[k] is not None:
                        ##如果请求数据中包含变量时需要做处理
                        if dict1[k].find(key) != -1:
                            dict1[k] = dict1[k].replace(key,value)
                            #
            one_test_data.append(dict1.copy())
        return one_test_data

    #更新初始化数据
    def update_init_data(self,sheetname,i,j):
        sh = self.wb.get_sheet_by_name(sheetname)
        init_data = self.get_init_variables(sheetname)
        sh.cell(i,j).value = str(int(init_data["${init_phone}"]) + 3)
        self.logger.info("更新后的初始化数据为：{0}".format(sh.cell(i,j).value))

    # def get_update_variables(self,sheetname,key,value):
    #     sh = self.wb.get_sheet_by_name(sheetname)
    #     list1 = []
    #     for i in range(2,sh.max_row+1):
    #         for j in range(1,sh.max_column):
    #             dict1 = {}
    #             key = sh.cell(i, 1)
    #             value = sh.cell(i, 2)
    #             dict1[sh.cell(i, 1)] = sh.cell(i, 2)
    #         list1.append(dict1.copy())
    def update_vari_data(self, sheetname, value,i, j):
        sh = self.wb.get_sheet_by_name(sheetname)
        sh.cell(i, j).value = value
        self.logger.info("更新后的动态变量为：{0}".format(sh.cell(i, j).value))

        #保存数据
    def savefile(self,filename):
        self.wb.save(filename)